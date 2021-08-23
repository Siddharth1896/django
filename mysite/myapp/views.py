from re import L
from django.http import response
from django.http.response import ResponseHeaders
from django.shortcuts import render
from django.http import HttpResponse, request
from flask.templating import render_template
from openpyxl.descriptors.base import Length
from openpyxl.workbook import workbook
from pandas.core.frame import DataFrame
from myapp.models import Data
from django.shortcuts import render, redirect
import pandas as pd
import openpyxl
import pandas as pd
from pandas import ExcelWriter
import numpy as np
from fuzzywuzzy import fuzz
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from werkzeug.utils import header_property, send_file
import xlsxwriter
import math
from os import link
from django.contrib.auth.forms import UserCreationForm
from .forms import CreateUserForm
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required


#### LOGIN & REGISTER LOGIC:

def registerPage(request):
    if request.user.is_authenticated:
        return redirect('myapp:dum')
    else:
        form = CreateUserForm()
        if request.method == 'POST':
            form = CreateUserForm(request.POST)
            if form.is_valid():
                form.save()
                user = form.cleaned_data.get('username')
                messages.success(request, 'Account was created for ' + user)
                return redirect('myapp:login')
        context = {'form':form}
        return render(request, 'register.html', context)

def loginPage(request):
    if request.user.is_authenticated:
        return redirect('myapp:dum')
    else:
        if request.method == 'POST':
            username = request.POST.get('username')
            password = request.POST.get('password')
            user = authenticate(request, username=username, password=password)
            if user is not None:
                login(request, user)
                return redirect('myapp:dum')
            else:
                messages.info(request, 'Username OR Password is Incorrect')
        context = {}
        return render(request, 'login.html', context)


def logoutUser(request):
    logout(request)
    return redirect('myapp:login')



#######
#######
# TO DISPLAY THE OUTPUT IN A NEW HTML PAGE
#######
#######
@login_required(login_url='myapp:login')
def dum(request):
    if "GET" == request.method:
        return render(request, 'dum.html', {})
    else:
        print(request.FILES['file'])
        file = request.FILES['file']
        new_ck = pd.read_excel(file)

        def formatNumber(num):
            if num % 1 == 0:
                return int(num)
            else:
                return num

### Part Number Match:
        df_data = pd.DataFrame(columns=['Date','Country','Vendor','Reference New Part Number','Part_number','License_name','Metric','Currency','Process_number','GLobal_price_listprice_USD','Discount_from_pricelist','Final_price_USD','Finalprice_of_Localcurrency','Awarded','Quantity','Volume_of_Deal'])
        for ind,row in new_ck.iterrows():
            num1 = row['Part_number']
            if type(num1) == str:
                num1 = num1
            else:
                num1 = formatNumber(num1)  

            rows = Data.objects.all().filter(Part_number = num1).values_list('Date_of_entry', 'Country', 'Vendor','Part_number','License_name','Metric','Currency','Process_number','GLobal_price_listprice_USD','Discount_from_pricelist','Final_price_USD','Finalprice_of_Localcurrency','Awarded','Quantity','Volume_of_Deal')
            for row in rows:
                for col_num in range(1):
                    a = pd.DataFrame([row], columns=['Date','Country','Vendor','Part_number','License_name','Metric','Currency','Process_number','GLobal_price_listprice_USD','Discount_from_pricelist','Final_price_USD','Finalprice_of_Localcurrency','Awarded','Quantity','Volume_of_Deal'])
                    a.insert(3, "Reference New Part Number", num1, True)                    
                    df_data=df_data.append(a)

        df_data1 = df_data.to_html(index=False)

### Part Name Match
        df_data2 = pd.DataFrame(columns=['Date','Country','Vendor','Reference New Part Number','Part_number','License_name','Metric','Currency','Process_number','GLobal_price_listprice_USD','Discount_from_pricelist','Final_price_USD','Finalprice_of_Localcurrency','Awarded','Quantity','Volume_of_Deal'])
        for ind,row in new_ck.iterrows():
            str1 = row['License_name']
            num1 = row['Part_number']
            if type(num1) == str:
                num1 = num1
            else:
                num1 = formatNumber(num1)
            
            direct_match = 'N'

            num2 = Data.objects.all().filter(Part_number = num1).values_list('Part_number')
            for row in num2:
                for col_num1 in range(len(row)):
                    if num1 == int(row[0]):
                        direct_match = 'Y'

            if direct_match == 'N':
                str2 = Data.objects.all().exclude(Part_number = num1).values_list('License_name')
                for row in str2:
                    for col_num1 in range(len(row)):
                        str22 = row[0]
                        r = fuzz.ratio(str(str1),str22)
                        if r > 50:
                            rows = Data.objects.all().filter(License_name = str22).values_list('Date_of_entry', 'Country', 'Vendor','Part_number','License_name','Metric','Currency','Process_number','GLobal_price_listprice_USD','Discount_from_pricelist','Final_price_USD','Finalprice_of_Localcurrency','Awarded','Quantity','Volume_of_Deal')
                            for row1 in rows:
                                for col_num1 in range(1):
                                    b = pd.DataFrame([row1], columns=['Date','Country','Vendor','Part_number','License_name','Metric','Currency','Process_number','GLobal_price_listprice_USD','Discount_from_pricelist','Final_price_USD','Finalprice_of_Localcurrency','Awarded','Quantity','Volume_of_Deal'])
                                    b.insert(3, "Reference New Part Number", num1, True)
                                    df_data2=df_data2.append(b)
        
        df_data3 = df_data2.to_html(index=False)

        return render(request, 'result.html',  {'df_data1': df_data1, 'df_data3': df_data3})


###########
###########
#   WORKS BASED ON A NEW OFFER FILE AND DATABASE - Use '/test' in Link
###########
###########
@login_required(login_url='myapp:login')
def test(request):
    
    if "GET" == request.method:
        return render(request, 'display.html', {})
    else:
        print(request.FILES['file'])
        file = request.FILES['file']
        new_ck = pd.read_excel(file)
        length = len(new_ck)
        new = pd.read_excel(file)

        def formatNumber(num):
            if num % 1 == 0:
                return int(num)
            else:
                return num

        w_book = xlsxwriter.Workbook('Data_result.xlsx')
        w_sheet = w_book.add_worksheet('Part_Number_Match')
        w_sheet1 = w_book.add_worksheet('Part_Name_Match')
        row_num = 3 + length 
        row_num1 = 0
        count = 0
        bold = w_book.add_format({'bold': True})


### PART NUMBER FETCH
        columns = ['Date','Country','Vendor','Part_number','License_name','Metric','Currency','Process_number','GLobal_price_listprice_USD','Discount_from_pricelist','Final_price_USD','Finalprice_of_Localcurrency','Awarded','Quantity','Volume_of_Deal']
        for col_num in range(len(columns)):
            w_sheet.write(row_num, col_num, columns[col_num], bold)
        for ind,row in new_ck.iterrows():
            num1 = row['Part_number']
            str1 = row['License_name']
            if type(num1) == str:
                num1 = num1
            else:
                num1 = formatNumber(num1)

            switch = 'N'
            dum_ref = 0
            rows = Data.objects.all().filter(Part_number = num1).values_list('Date_of_entry', 'Country', 'Vendor','Part_number','License_name','Metric','Currency','Process_number','GLobal_price_listprice_USD','Discount_from_pricelist','Final_price_USD','Finalprice_of_Localcurrency','Awarded','Quantity','Volume_of_Deal')
            for row in rows:
                row_num += 1
                if dum_ref == 0:
                    new.loc[ind,"More Details"] = row_num + 1
                    dum_ref = dum_ref+1

                for col_num in range(len(row)):
                    w_sheet.write(row_num, col_num, row[col_num])
                    switch = 'Y'
            
            if switch == 'Y':
                new_ck.loc[ind,"Exact Match"] = 'Yes'
            elif np.isnan(num1):
                new_ck.loc[ind,"Exact Match"] = 'No'
            else:
                new_ck.loc[ind,"Exact Match"] = 'FUZZY'
        
### PART NAME FETCH
            if switch == 'N':
                if count == 0:
                    df_data2 = pd.DataFrame(columns=['Date','Country','Vendor','Reference New Part Number','Part_number','License_name','Metric','Currency','Process_number','GLobal_price_listprice_USD','Discount_from_pricelist','Final_price_USD','Finalprice_of_Localcurrency','Awarded','Quantity','Volume_of_Deal'])
                    columns = ['Date','Country','Vendor','Part_number','License_name','Metric','Currency','Process_number','GLobal_price_listprice_USD','Discount_from_pricelist','Final_price_USD','Finalprice_of_Localcurrency','Awarded','Quantity','Volume_of_Deal']
                    count = count + 1
                    for col_num in range(len(columns)):
                        w_sheet1.write(row_num1, col_num, columns[col_num], bold)
                str2 = Data.objects.all().values_list('License_name')
                for row in str2:
                    for col_num2 in range(len(row)):
                        str22 = row[0]
                        r = fuzz.ratio(str(str1),str22)
                        if r > 50:
                            rows = Data.objects.all().filter(License_name = str22).values_list('Date_of_entry', 'Country', 'Vendor','Part_number','License_name','Metric','Currency','Process_number','GLobal_price_listprice_USD','Discount_from_pricelist','Final_price_USD','Finalprice_of_Localcurrency','Awarded','Quantity','Volume_of_Deal')
                            for row1 in rows:
                                row_num1 += 1
                                for col_num2 in range(len(row1)):
                                    w_sheet1.write(row_num1, col_num, row1[col_num])
                                for col_num2 in range(1):
                                    b = pd.DataFrame([row1], columns=['Date','Country','Vendor','Part_number','License_name','Metric','Currency','Process_number','GLobal_price_listprice_USD','Discount_from_pricelist','Final_price_USD','Finalprice_of_Localcurrency','Awarded','Quantity','Volume_of_Deal'])
                                    b.insert(3, "Reference New Part Number", num1, True)
                                    df_data2=df_data2.append(b)
    
              
        w_book.close()

        book = load_workbook('Data_result.xlsx')
        writer1 = pd.ExcelWriter('Data_result.xlsx', sheetname='Part_Number_Match', engine='openpyxl')
        writer1.book = book
        writer1.sheets = {ws.title: ws for ws in book.worksheets}

        for sheetname in writer1.sheets:
            new_ck.to_excel(writer1,sheet_name='Part_Number_Match', startrow=0, index = False,header= True)
        writer1.save()

        book = load_workbook('Data_result.xlsx')
        writer1 = pd.ExcelWriter('Data_result.xlsx', sheetname='Part_Name_Match', engine='openpyxl')
        writer1.book = book
        writer1.sheets = {ws.title: ws for ws in book.worksheets}

        for sheetname in writer1.sheets:
            df_data2.to_excel(writer1,sheet_name='Part_Name_Match', startrow=0, index = False,header= True)
        writer1.save()

### Pattern Fill:

        indexy = new.index
        rowse = len(indexy)
        w_color = load_workbook('Data_result.xlsx')
        w_sheet = w_color['Part_Number_Match']
        w_sheet1 = w_color['Part_Name_Match']
        fill_green = PatternFill(patternType='solid', fgColor='14F232')
        fill_orange = PatternFill(patternType='solid', fgColor='E2930A')
        fill_red = PatternFill(patternType='solid', fgColor='D82323')
        fill_blue = PatternFill(patternType='solid', fgColor='25AAEC')
        color_sw = 'NA'
        county = 1

        for ind7,row in new.iterrows():
            num11 = row['Part_number']
            if type(num11) == str:
                num11 = num11
            else:
                num11 = formatNumber(num11)

            new_price = row['Discounted Unitary Price']
            pointer = row['More Details']
            if pointer != 'NA' and pointer == pointer:
                pointer = int(pointer)

            old_price = Data.objects.all().filter(Part_number = num11).values_list('GLobal_price_listprice_USD')
            for row in old_price:
                for col_num1 in range(len(row)):
                    per = ((new_price - row[0])/row[0])*100
                    if per < 0:
                        color_sw = 'G'
                    if per > 0 and per <20 :
                        color_sw = 'O'
                    if per > 20:
                        color_sw = 'R'
            leny = len(new.columns)
            for row in w_sheet.iter_rows(min_row=1, max_row = rowse, min_col=leny, max_col=leny):
                for cell in row:
                    if cell.value == 'Yes' and color_sw != 'NA':
                        if color_sw == 'G':
                            cell.fill = fill_green
                            link = "#Part_Number_Match!D"+str(pointer)
                            cell.value = '=HYPERLINK("{}", "{}")'.format(link, "Yes")
                            color_sw = 'NA'
                            w_color.save('Data_result.xlsx')
                            break                                            
                        if color_sw == 'O':
                            cell.fill = fill_orange
                            link = "#Part_Number_Match!D"+str(pointer)
                            cell.value = '=HYPERLINK("{}", "{}")'.format(link, "Yes")
                            color_sw = 'NA'
                            w_color.save('Data_result.xlsx')
                            break
                        if color_sw == 'R':
                            cell.fill = fill_red
                            link = "#Part_Number_Match!D"+str(pointer)
                            cell.value = '=HYPERLINK("{}", "{}")'.format(link, "Yes")
                            color_sw = 'NA'
                            w_color.save('Data_result.xlsx')
                            break
            first_time = 'Y'
            for row in w_sheet.iter_rows(min_row=1, max_row = rowse+1, min_col=leny, max_col=leny):
                for cell in row:
                    if cell.value == 'FUZZY' and first_time == 'Y':
                        for ind4,row in df_data2.iterrows():
                            num111 = row['Reference New Part Number']

                            if num11 == num111 and first_time == 'Y':
                                county = county + 1
                                first_time = 'N'
                                cell.fill = fill_blue
                                link = "#Part_Name_Match!D"+str(county)
                                cell.value = '=HYPERLINK("{}", "{}")'.format(link, "FUZZY")
                                w_color.save('Data_result.xlsx')
                            elif num11 == num111 and first_time == 'N':
                                county = county + 1

        w_color.save('Data_result.xlsx')

        with open('Data_result.xlsx', 'rb') as model_excel:
            result = model_excel.read()
        response = HttpResponse(result)
        response['Content-Disposition'] = 'attachment; filename=Data_result.xlsx'
        return response


##########
##########
##########
##########
# WORKS BASED ON A REFERENCE FILE AND NEW OFFER EXCEL FILE - Use '/' in Link
##########
##########
##########
@login_required(login_url='myapp:login')
def index(request):
    if "GET" == request.method:
        return render(request, 'start.html', {})

@login_required(login_url='myapp:login')
def number(request):
    if "GET" == request.method:
        return render(request, 'process.html', {})

    else:
##############################
# 1 # Input file fetch:
##############################

        print(request.FILES['file'])
        f = request.FILES['file']
        ref = pd.read_excel(f, sheet_name='SW_Opcional_On premise', skiprows = 5)
        ref = ref.drop(['Unnamed: 0'], axis = 1)
        ref_name = ref
        df11 = ref
        ref_name.to_excel("Reference_name.xlsx", index=False)

        print(request.FILES['file1'])
        f1 = request.FILES['file1']
        new = pd.read_excel(f1)
        new_name = new
        df22 = new
        new_name.to_excel("New_Offer_name.xlsx", index=False)

    ##############################
    # 2 # Logic for Part Number Match:
    ##############################

        res = pd.merge(ref, new, on="Part_number")
        print(res)

        res.rename(columns = {'Discounted Unitary Price_y':'New Discounted Unitary Price','Discounted Unitary Price_x':'Old Discounted Unitary Price'},inplace = True)
        res.rename(columns = {'License_name_y':'New License_name','License_name_x':'Old License_name'},inplace = True)
        print(res)
        final = res[['Part_number','New License_name','Old License_name','Metric','Country','Company / OB','Total Price per Product ','Total Purchase Price','Old Discounted Unitary Price','New Discounted Unitary Price']].dropna()

    ##############################
    # 3 # Calculated fields Difference, Percentage Variation, Deal - Yes/No:
    ##############################

        for ind1,row in final.iterrows():
            final.loc[ind1,"Difference"] = row['New Discounted Unitary Price'] - row['Old Discounted Unitary Price']
            final.loc[ind1,"Percentage Variation"] = ((row['New Discounted Unitary Price'] - row['Old Discounted Unitary Price']) / row['Old Discounted Unitary Price']) * 100

        for ind2,row in final.iterrows():
            if row['Percentage Variation'] <= 5:
                final.loc[ind2,"Deal - Yes/No"] = 'Yes'
            else:
                final.loc[ind2,"Deal - Yes/No"] = 'No'

    ##############################
    # 4 # Add Reference fields - Found???, Ref Count, Best Reference Price, Several References
    ##############################

        for ind3,row in new.iterrows():
            num1 = row['Part_number']
            switch = 'N'
            count = 0

            for ind,row in final.iterrows():
                if row['Part_number'] == num1:
                    switch = 'Y'
                    count = count + 1

            if switch == 'Y':
                new.loc[ind3,"Exact Match"] = 'Yes'
            elif np.isnan(num1):
                new.loc[ind3,"Exact Match"] = 'No'
            else:
                new.loc[ind3,"Exact Match"] = 'FUZZY'

        new.to_excel("Master_Result.xlsx", index=False)

        book = load_workbook('Master_Result.xlsx')
        writer1 = pd.ExcelWriter('Master_Result.xlsx', sheetname='Sheet1', engine='openpyxl')
        writer1.book = book
        writer1.sheets = {ws.title: ws for ws in book.worksheets}

        for sheetname in writer1.sheets:
            counter = writer1.sheets[sheetname].max_row + 7
            final.to_excel(writer1,sheet_name='Sheet1', startrow=writer1.sheets[sheetname].max_row + 5, index = False,header= True)

            for ind4,row1 in final.iterrows():
                final.loc[ind4,"More Details"] = counter
                counter = counter + 1
        writer1.save()
        writer1.close() 


    ##############################
    # 5 # Add Reference field - More Details
    ##############################

        for ind5,row in new.iterrows():
            num11 = row['Part_number']
            check = 'N'
            count1 = 0

            for ind6,row in final.iterrows():
                num22 = row['Part_number']
                if num22 == num11:
                    count1 = count1 + 1
                if num22 == num11 and count1 == 1:
                    check = 'Y'
                    more = row['More Details']

            if check == 'Y':
                new.loc[ind5,"More Details"] = more
            else:
                new.loc[ind5,"More Details"] = 'NA'

    ##############################
    # DO NOT DELETE - 6 # Write the output to Master_Result Excel file
    ##############################

        # book = load_workbook('Master_Result.xlsx')
        # writer2 = pd.ExcelWriter('Master_Result.xlsx', sheetname='Sheet1', engine='openpyxl')
        # writer2.book = book
        # writer2.sheets = {ws.title: ws for ws in book.worksheets}

        # for sheetname in writer2.sheets:
        #     new.to_excel(writer2,sheet_name='Sheet1', startrow=0, index = False,header= True)
        # writer2.save()
        # writer2.close()

    ##############################
    # 7 # Pattern FIll
    ##############################
        
        indexy = new.index
        rows = len(indexy)
        rows = rows + 1

        w_color = load_workbook('Master_Result.xlsx')
        w_sheet = w_color['Sheet1']
        
        fill_green = PatternFill(patternType='solid', fgColor='14F232')
        fill_orange = PatternFill(patternType='solid', fgColor='E2930A')
        fill_red = PatternFill(patternType='solid', fgColor='D82323')
        fill_blue = PatternFill(patternType='solid', fgColor='25AAEC')

        color_sw = 'NA'

        row_count = 1

        for ind7,row in new.iterrows():
            num222 = row['Part_number']
            pointer = row['More Details']
            if pointer != 'NA':
                pointer = int(pointer)
            yn = row['Exact Match']
            switcher = 'N'
            row_count = row_count + 1

            for ind8,row in final.iterrows():
                num111 = row['Part_number']
                per = row['Percentage Variation']
                switcher = 'N'

                if num111 == num222:
                    switcher = 'Y'
                
                if switcher == 'Y' and yn == 'Yes' and per < 0:
                    color_sw = 'G'

                if switcher == 'Y' and yn == 'Yes' and per > 0 and per <20 :
                    color_sw = 'O'

                if switcher == 'Y' and yn == 'Yes' and per > 20:
                    color_sw = 'R'

                if switcher == 'Y': 
                    for row in w_sheet.iter_rows(min_row = 2, max_row = rows, min_col=2, max_col=2):

                        for cell in row:
                            if cell.value == num222:

                                for row in w_sheet.iter_rows(min_row=row_count, max_row = row_count, min_col=12, max_col=12):
                                    for cell in row:
                                        if cell.value == 'Yes' and color_sw != 'NA':
                                            if color_sw == 'G':
                                                cell.fill = fill_green
                                                link = "#Sheet1!A"+str(pointer)
                                                cell.value = '=HYPERLINK("{}", "{}")'.format(link, "Yes")
                                                color_sw = 'NA'
                                                w_color.save('Master_Result.xlsx')
                                                break                                            
                                            if color_sw == 'O':
                                                cell.fill = fill_orange
                                                link = "#Sheet1!A"+str(pointer)
                                                cell.value = '=HYPERLINK("{}", "{}")'.format(link, "Yes")
                                                color_sw = 'NA'
                                                w_color.save('Master_Result.xlsx')
                                                break
                                            if color_sw == 'R':
                                                cell.fill = fill_red
                                                link = "#Sheet1!A"+str(pointer)
                                                cell.value = '=HYPERLINK("{}", "{}")'.format(link, "Yes")
                                                color_sw = 'NA'
                                                w_color.save('Master_Result.xlsx')
                                                break

        w_color.save('Master_Result.xlsx')

    ################################
    # 8 # Logic for Part Name Match & Shading in Sheet 1 for 'FUZZY':
    ################################

        book = load_workbook('Master_Result.xlsx')
        writer3 = pd.ExcelWriter('Master_Result.xlsx', engine = 'openpyxl')
        writer3.book = book
        secondMockData = { ' ': []}
        secondMockDF = pd.DataFrame(secondMockData)
        secondMockDF.to_excel(writer3, sheet_name = 'Name_Match')
        writer3.save()

        pointer1 = 0
        rowIndex = 0

        for ind7,row in new.iterrows():
            
            pnum1 = row['Part_number']
            str2 = row['License_name']
            no_match_sw = 'Y'

            for ind8,row in final.iterrows():
                pnum2 = row['Part_number']
                if pnum1 == pnum2 or np.isnan(pnum1):
                    no_match_sw = 'N'

            if no_match_sw == 'Y':

                    wb = openpyxl.load_workbook('Reference_name.xlsx')
                    wb.sheetnames
                    a = wb["Sheet1"]

                    for row in a.iter_rows(min_col=12,max_col=12,min_row=2):
                        for cell in row:
                            str1 = cell.value

                            r = fuzz.ratio(str1,str2)

                            if r > 50:

                                res11 = df11[['Part_number','License_name','Discounted Unitary Price','Licensing Metric','Country','Company / OB','Total Price per Product ','Total Purchase Price']].where(df11['License_name'] == str1).dropna()
                                res22 = df22[['Part_number','License_name','Discounted Unitary Price']].where(df22['License_name'] == str2).dropna()

                                res13 = res11.rename(columns = {'Part_number':'Ref Part_number','Discounted Unitary Price':'Ref Discounted Unitary Price','License_name':'Ref License_name'},inplace = False)
                                res23 = res22.rename(columns = {'Part_number':'New Part_number','Discounted Unitary Price':'New Discounted Unitary Price','License_name':'New License_name'},inplace = False)
                                
                                res13 = res13.reset_index(drop=True)
                                res23 = res23.reset_index(drop=True)

                                final_name = pd.concat([res23,res13],axis=1)

                                checker = 0

                                if rowIndex == 0:
                                    book = load_workbook('Master_Result.xlsx')
                                    writer4 = pd.ExcelWriter('Master_Result.xlsx', sheetname='Name_Match',engine='openpyxl')
                                    writer4.book = book
                                    writer4.sheets = {ws.title: ws for ws in book.worksheets}
                                    for sheetname in writer4.sheets:
                                            final_name.to_excel(writer4,sheet_name='Name_Match', startrow=0, index = False,header= True)
                                            rowIndex = rowIndex + 1
                                            pointer1 = pointer1 + 1
                                    writer4.save()
                                else:
                                    book = load_workbook('Master_Result.xlsx')
                                    writer4 = pd.ExcelWriter('Master_Result.xlsx', sheetname='Name_Match',engine='openpyxl')
                                    writer4.book = book
                                    writer4.sheets = {ws.title: ws for ws in book.worksheets}
                                    for sheetname in writer4.sheets:
                                        if checker == 0:
                                            final_name.to_excel(writer4,sheet_name='Name_Match', startrow=rowIndex, index = False, header=False)
                                            rowIndex = rowIndex + 1
                                            checker = checker + 1
                                            pointer1 = pointer1 + 1
                                    writer4.save()

                                checky = 1
                                row_counter = 0

                                for ind8,row in new.iterrows():
                                    pnum22 = row['Part_number']

                                    if pnum1 == pnum22:
                                        row_counter = row_counter + 1
                                        checky = checky + row_counter
                                    else:
                                        row_counter = row_counter + 1
                                    
                                    if checky != 1:    
                                        w_color = load_workbook('Master_Result.xlsx')
                                        w_sheet = w_color['Sheet1']
                                        for row in w_sheet.iter_rows(min_row = checky,max_row=checky, min_col=12, max_col=12):
                                            for cell in row:
                                                if cell.value == 'FUZZY' and checky != 1:
                                                    cell.fill = fill_blue
                                                    link = "#Name_Match!A"+str(pointer1)
                                                    cell.value = '=HYPERLINK("{}", "{}")'.format(link, "FUZZY")
                                                    checky = 1
                                        w_color.save('Master_Result.xlsx')

                    book.close()  



        return render(request,'downloadnumber.html',{})

@login_required(login_url='myapp:login')
def downloadnumber_file(request):
    with open('Master_Result.xlsx', 'rb') as model_excel:
        result = model_excel.read()
    response = HttpResponse(result)
    response['Content-Disposition'] = 'attachment; filename=Master_Result.xlsx'
    return response
